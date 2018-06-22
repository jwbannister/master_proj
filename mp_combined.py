#! /usr/bin/env python
import pandas as pd
import numpy as np
import datetime
import os
import sys
from collections import OrderedDict
from openpyxl import worksheet
from openpyxl import load_workbook
from itertools import compress

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/364/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A0:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/364/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A0:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+0)
        # cols = range(min_col, max_col+0)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 0, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
patch_worksheet()

def build_factor_tables():
    dcm_factors = mp_file.parse(sheet_name="Generic HV & WD", header=2, \
            usecols="A,C,D,E,F,G,H", \
            names=["dcm", "bw", "mw", "pl", "ms", "md", "water"])[0:31]
    dcm_factors.set_index('dcm', inplace=True)
    # build up custom habitat and water factor tables
    custom_info = mp_file.parse(sheet_name="Custom HV & WD", header=0, \
            usecols="A,B,C,D,E,F,G,H,I", \
            names=["dca", "dcm", "step", "bw", "mw", "pl", "ms", "md", "water"])
    custom_info.set_index(['dca', 'dcm'], inplace=True)
    custom_filled = custom_info.apply(backfill, axis=1, \
            backfill_data=dcm_factors, \
            columns_list=['bw', 'mw', 'pl', 'ms', 'md', 'water'])
    custom_steps = ['base', 'dwm', 'step0', 'mp']
    factors = {x: build_custom_steps(x, custom_steps, custom_filled) \
            for x in custom_steps}
    factors['dcm'] = dcm_factors.copy()
    return factors

def read_dca_info():
    dca_info = mp_file.parse(sheet_name="MP_new", header=None, skiprows=21, \
            usecols="A,B,D,F", \
            names=["dca", "area_ac", "base", "step0"])
    dca_info['area_sqmi'] = dca_info['area_ac'] * 0.0015625
    dca_info.set_index('dca', inplace=True)
    return dca_info

def read_past_status():
    lake_case = {'base': [], 'step0': []}
    for case in lake_case.keys():
        assignments = [np.array(factors['dcm'].index.get_level_values('dcm')) \
                == dca_info[case][x] for x in range(0, len(dca_info))]
        assignments = [assignments[x].astype(int).tolist() \
                for x in range(0, len(assignments))]
        case_df = pd.DataFrame(assignments)
        case_df.index = dca_info.index
        case_df.columns = factors['dcm'].index.tolist()
        lake_case[case] = case_df
    return lake_case

def define_soft():
    soft_dcm_input = mp_file.parse(sheet_name="MP Analysis Input", header=6, \
            usecols="A").iloc[:, 0].tolist()[5:12]
    soft_dcms = [x for x in soft_dcm_input if x in dcm_list]
    soft_idx = [x for x, y in enumerate(factors['dcm'].index.tolist()) if y in soft_dcms]
    return soft_idx

def get_area(dca_case, dca_name, custom_factors, dca_info, hab):
    dcm_name = dcm_list[dca_case.index(1)]
    if (dca_name, dcm_name) in [(x, y) for x, y in custom_factors['mp'].index.tolist()]:
       case_factors = custom_factors['mp'].loc[dca_name, dcm_name]
    else:
       case_factors = custom_factors['dcm'].iloc[dca_case.index(1)]
    area = case_factors * dca_info.loc[dca_name]['area_sqmi']
    return area[hab]

def evaluate_dca_change(dca_case, previous_case, previous_factors, custom_factors, \
        priority, dca_name, dca_info, waterless_preferences):
    previous_case_factors = previous_factors.loc[previous_case.name]
    dcm_name = custom_factors['dcm'].index.tolist()[dca_case.index(1)]
    if (dca_name, dcm_name) in [(x, y) for x, y in custom_factors['mp'].index.tolist()]:
       case_factors = custom_factors['mp'].loc[dca_name, dcm_name]
    else:
       case_factors = custom_factors['dcm'].iloc[dca_case.index(1)]
    if priority[1]=='water':
        smart = case_factors['water'] - previous_case_factors['water'] < 0
        benefit1 = previous_case_factors['water'] - case_factors['water']
        try:
            benefit2 = waterless_preferences[dcm_name]
        except:
            benefit2 = -6
    else:
        p1_increase = case_factors[priority[1]] - previous_case_factors[priority[1]]
        water_increase = 100 + (case_factors['water'] - previous_case_factors['water'])
        smart = p1_increase > 0
        benefit1 = p1_increase / water_increase
        p2_increase = case_factors[priority[2]] - previous_case_factors[priority[2]]
        benefit2 = p2_increase / water_increase
    return {'smart': smart, 'benefit1':benefit1, 'benefit2':benefit2}

def prioritize(value_percents, hab_minimums):
    hab_deficits = {x: value_percents[x] - hab_minimums[x] \
            for x in value_percents.index.tolist() \
            if x != 'water'}
    if any([x < 0 for x in hab_deficits.values()]):
        sort_deficits = sorted(hab_deficits.values())
        one = hab_deficits.values().index(sort_deficits[0])
        two = hab_deficits.values().index(sort_deficits[1])
        return {1: hab_deficits.keys()[one], 2: hab_deficits.keys()[two]}
    else:
        return {1: 'water', 2: value_percents[0:5].idxmin()}

def backfill(row, backfill_data, columns_list):
    for col in columns_list:
        if np.isnan(row[col]):
            row[col] = backfill_data.loc[row.name[1], col]
    return row

def build_custom_steps(step, step_list, data):
    factors = data.loc[data['step']==step_list[0], :].copy()
    sub_steps = [step_list[x] for x in range(1, step_list.index(step)+1)]
    for sub in sub_steps:
        sub_df = data.loc[data['step']==sub, :].copy()
        for idx in factors.index:
            if idx in sub_df.index:
                factors.drop(idx, inplace=True)
        factors = factors.append(sub_df)
    factors.drop('step', axis=1, inplace=True)
    return factors

def get_assignments(case, dca_list, dcm_list):
    assignments = pd.DataFrame([dcm_list[row.tolist().index(1)] \
            for index, row in case.iterrows()], index=dca_list, columns=['dcm'])
    return assignments

def build_case_factors(lake_case, custom_factors, stp):
    factors = pd.DataFrame()
    for idx in range(0, len(lake_case)):
        dca_name = lake_case.iloc[idx].name
        dca_case = lake_case.iloc[idx]
        dcm_name = dca_case[dca_case==1].index[0]
        dca_idx = [x for x, y in \
                enumerate(custom_factors[stp].index.get_level_values('dca')) \
                if y==dca_name]
        dcm_idx = [x for x, y in \
                enumerate(custom_factors[stp].index.get_level_values('dcm')) \
                if y==dcm_name]
        custom_idx = [x for x in dca_idx if x in dcm_idx]
        if len(custom_idx)>0:
            tmp = custom_factors[stp].iloc[custom_idx].copy()
            tmp['dca'] = lake_case.index.tolist()[idx]
            factors = factors.append(tmp)
        else:
            tmp = custom_factors['dcm'].loc[dcm_name].copy()
            tmp['dca'] = lake_case.index.tolist()[idx]
            factors = factors.append(tmp)
    factors.set_index('dca', inplace=True)
    return factors

def build_single_case_factors(dca_case, dca_name, custom_factors, stp):
    factors = pd.DataFrame()
    dcm_name = dcm_list[dca_case.index(1)]
    try:
        factors = custom_factors[stp].loc[(dca_name, dcm_name)]
    except:
        factors = custom_factors['dcm'].loc[dcm_name]
    return factors

def calc_totals(case, custom_factors, step, dca_info):
    dca_list = dca_info.index.tolist()
    factors = build_case_factors(case, custom_factors, step)
    return factors.multiply(np.array(dca_info['area_ac']), axis=0).sum()

def printout(flag):
    readout = ""
    if flag == 'screen':
        for x in new_percent.keys().tolist():
            if priority[1] == x:
                pri = color.BOLD
            else:
                pri = ""
            try:
                if new_percent[x] >= hab_limits[x]:
                    readout = readout + pri + color.GREEN + x + ": " + \
                            str(round(new_percent[x], 3)) \
                            + ", " + color.END
                else:
                    readout = readout + pri + color.RED + x + ": " + \
                            str(round(new_percent[x], 3)) \
                            + ", " + color.END
            except:
                readout = readout + pri + x + ": " + \
                        str(round(new_percent[x], 3)) \
                        + ", " + color.END
        return readout
    else:
        for x in new_percent.keys().tolist():
            readout = readout + x + ": " + \
                    str(round(new_percent[x], 3)) \
                    + ", "
        readout = readout + "priority1 = " + priority[1] + ", priority2 = " +\
                priority[2]
    return readout

def check_exceed_area(test_totals, new_totals, limits, guild_available):
    exceed_flag = {x: 'ok' for x in guild_list if x != 'water'}
    for x in exceed_flag.keys():
        if (test_totals[x] + (guild_available[x] * 640)) / total['base'][x] < limits[x]:
            exceed_flag[x] = 'under'
    # meadow is hard to establish, do not want to reduce only to have to
    # re-establish. Prevent meadow from dipping below target value and never
    # add any more meadow
    if not unconstrained_case:
        if test_totals['md'] / total['base']['md'] < limits['md']:
            exceed_flag['md'] = 'under'
        if test_totals['md'] > new_totals['md'] :
            exceed_flag['md'] = 'over'
    return exceed_flag

def set_constraint(axis, idx, new_constraint, constraint_df):
    new = new_constraint
    if axis == 1:
        existing = constraint_df.loc[idx].tolist()
        constraint_df.loc[idx] = [min(x, y) for x, y in zip(new, existing)]
    else:
        existing = constraint_df[idx]
        constraint_df[idx] = [min(x, y) for x, y in zip(new, existing)]
    return constraint_df

def get_guild_available(smart_cases):
    guild_available = {}
    available_hard = hard_limit - hard_transition
    for hab in guild_list:
        temp = []
        for dca in set([x[3] for x in smart_cases[1:]]):
            try:
                temp.append(\
                        sorted([[x[5][hab], dca_info.loc[dca]['area_sqmi']] \
                        for x in smart_cases if x[3] == dca and \
                        dca_info.loc[dca]['area_sqmi'] < available_hard], \
                        reverse=True)[0])
            except:
                temp.append([0, 0])
        temp = [x for x in sorted(temp, reverse=True) if x[1] < available_hard]
        while sum([x[1] for x in temp]) > 0.8 * available_hard:
            temp.pop()
        guild_available[hab] = sum([x[0] for x in temp])
    return guild_available

# set option flags
unconstrained_case = True
freeze_farm = False
factor_water = True
preset_base_water = 73351
file_flag = ""
if unconstrained_case: file_flag = file_flag + " NO_CONSTRAINTS"
if freeze_farm: file_flag = file_flag + " FARM_FROZEN"
if factor_water: file_flag = file_flag + " H20_ADJUST"

# read data from original Master Project planning workbook
file_path = os.path.realpath(os.getcwd()) + "/"
file_name = "MP LAUNCHPAD.xlsx"
mp_file = pd.ExcelFile(file_path + file_name)
timestamp = datetime.datetime.now().strftime('%m_%d_%y %H_%M')
output_log = file_path + "output/" + "MP " + "LOG " + timestamp + file_flag + '.txt'
output_excel = file_path + "output/" + "MP " + timestamp + file_flag + '.xlsx'
output_csv = file_path + "output/mp_steps " + timestamp + file_flag + '.csv'
log_file = open(output_log, 'a')

factors = build_factor_tables()
dca_info = read_dca_info()
dca_list = dca_info.index.tolist()
dcm_list = factors['dcm'].index.tolist()
hab_terms = ['BWF', 'MWF', 'SNPL', 'MSB', 'Meadow']
hdcm_identify = [any([y in x for y in hab_terms]) for x in dcm_list]
hdcm_list = list(compress(dcm_list, hdcm_identify))
guild_list = [x for x in factors['dcm'].columns if x != 'water']

# read and set preferences for waterless DCMs
waterless_input = mp_file.parse(sheet_name="MP Analysis Input", header=19, \
        usecols="A").iloc[:, 0].tolist()
waterless_dict = {x:-y for x, y in zip(waterless_input, range(1, 6))}

# generate list of "soft transition" DCMS
soft_idx = define_soft()

# read limits and toggles
script_input = mp_file.parse(sheet_name="MP Analysis Input", header=None, \
        usecols="B").iloc[:, 0].tolist()[0:4]
hard_limit = script_input[0]
soft_limit = script_input[1]
dcm_limits = {'Brine': script_input[2]}
hab_limit_input = mp_file.parse(sheet_name="MP Analysis Input", \
        usecols="B,C").iloc[6:11, 0:2]
hab_limits = dict(zip(hab_limit_input.iloc[:, 1].tolist(), \
        hab_limit_input.iloc[:, 0].tolist()))

# initialize constraints tables
start_constraints = pd.DataFrame(np.ones((len(dca_list), len(dcm_list)), \
        dtype=np.int8))
start_constraints.index = dca_list
start_constraints.columns = dcm_list
step_constraints = pd.DataFrame(np.ones((len(dca_list), 5), dtype=np.int8))
step_constraints.index = dca_list
step_constraints.columns = range(1, 6)

if not unconstrained_case:
    # set hard-wired constraints
    # Constraint to only remove "Meadow" habitat is wired into
    # check_exceed_area function
    # nothing allowed as Enhanced Natural Vegetation (ENV) except Channel Areas
    new = [1 if 'Channel' in x else 0 for x in dca_list]
    start_constraints = set_constraint(0, 'ENV', new, start_constraints)
    # Channel areas to remain unchanged
    for dca in ['Channel Area North', 'Channel Area South']:
        new = [1 if x == 'ENV' else 0 for x in dcm_list]
        start_constraints = set_constraint(1, dca, new, start_constraints)
    # no additional sand fences besides existing T1A1
    new = [1 if x == 'T1A-1' else 0 for x in dca_list]
    start_constraints = set_constraint(0, 'Sand Fences', new, start_constraints)
    # DCAs currently designated as HDCM to remain unchanged
    current_hab = dca_info.loc[[x in hdcm_list for x in dca_info['step0']]]
    for dca in current_hab.index:
        new = [1 if x == current_hab.loc[dca]['step0'] else 0 for x in dcm_list]
        start_constraints = set_constraint(1, dca, new, start_constraints)
    # all DCAs currently under waterless DCM should remain unchanged
    waterless = dca_info.loc[[x in waterless_dict.keys() for x in dca_info['step0']]]
    for dca in waterless.index:
        new = [1 if x == waterless.loc[dca]['step0'] else 0 for x in dcm_list]
        start_constraints = set_constraint(1, dca, new, start_constraints)
    # all DCAs under dust control, no "None" DCMs allowed
    new = [0 for x in dca_list]
    start_constraints = set_constraint(0, 'None', new, start_constraints)

    # read in and implement Ops constraints from LAUNCHPAD
    constraints_input = mp_file.parse(sheet_name="Constraints", header=11, \
            usecols="A:N")
    constraints_input.rename(columns={'DCM Constraints': 'dca'}, inplace=True)
    constraints_input.set_index('dca', inplace=True)
    step_start = ['Step' in str(x) for \
            x in constraints_input.index.tolist()].index(True)
    dcm_constraint_input = constraints_input.iloc[:step_start, :].dropna(how='all')
    step_constraint_input = constraints_input.iloc[step_start+1:, :].dropna(how='all')
    for dca in dcm_constraint_input.index:
        not_allowed = dcm_constraint_input.loc[dca].dropna().tolist()
        for dcm in not_allowed:
            new = [0 if x == dcm else 1 for x in dcm_list]
            start_constraints = set_constraint(1, dca, new, start_constraints)
    for dca in step_constraint_input.index:
        not_allowed = step_constraint_input.loc[dca].dropna().tolist()
        for step in not_allowed:
            new = [0 if x == step else 1 for x in range(1, 6)]
            step_constraints = set_constraint(1, dca, new, step_constraints)

if freeze_farm:
    farm_dcas = dca_info.loc[[x == 'Veg 08' for x in dca_info['step0']]]
    for dca in farm_dcas.index:
        new = [1 if x == 'Veg 08' else 0 for x in dcm_list]
        start_constraints = set_constraint(1, dca, new, start_constraints)

# read in past DCA/DCM assignments from LAUNCHPAD
lake_case = read_past_status()
if factor_water:
    calc_base_water = calc_totals(lake_case['base'], factors, 'base', dca_info).loc['water']
    water_adjust = preset_base_water/calc_base_water
    for step in factors.keys():
        factors[step]['water'] = water_adjust * factors[step]['water']
total = {}
for case in lake_case.keys():
    total[case] = calc_totals(lake_case[case], factors, case, dca_info)

# initialize variables before loop
constraints = start_constraints.copy()
case = lake_case['step0'].copy()
assignments = get_assignments(case, dca_list, dcm_list)
assignments.columns = ["step0"]
case_factors = build_case_factors(case, factors, 'mp')
new_percent = total['step0']/total['base']
new_total = total['step0'].copy()
priority = prioritize(new_percent, hab_limits)

step_info = {}
step_info['base'] = {'totals': total['base'],
        'percent_base': total['base']/total['base'], \
        'hard_transition': 0, 'soft_transition': 0, \
        'assignments': get_assignments(lake_case['base'], dca_list, dcm_list)}
step_info[0] = {'totals': new_total, 'percent_base': new_percent, \
        'hard_transition': 0, 'soft_transition': 0, \
        'assignments': assignments}

dca_water = pd.DataFrame({'step0': case_factors['water'].multiply(dca_info['area_ac'], \
        axis=0)})

tracking = pd.DataFrame.from_dict({'dca': [], 'mp': [], 'step': []})

change_counter = 0
for step in range(1, 6):
    # intialize step area limits
    hard_transition, soft_transition = 0, 0
    dcm_area_tracking = {'Brine': 0}
    while hard_transition < hard_limit or soft_transition < soft_limit:
        change_counter += 1
        output = "step " + str(step) + ", change " + str(change_counter) + \
                ": hard/soft " + str(round(hard_transition, 2)) + "/" + \
                str(round(soft_transition, 2))
        print output
        print printout('screen')
        log_file.write(output + "\n")
        log_file.write(printout('log') + "\n")
        smart_cases = []
        for dca in dca_list:
            if step_constraints.loc[dca, step] != 0:
                tmp = constraints.loc[dca].tolist()
                tmp_ind = [x for x, y in enumerate(tmp) if y == 1]
                for dcm_ind in tmp_ind:
                    if dcm_ind in soft_idx:
                        flag = 'soft'
                    else:
                        flag='hard'
                    b = [0 for x in tmp]
                    b[dcm_ind] = 1
                    case_eval = evaluate_dca_change(b, case.loc[dca], case_factors, \
                            factors, priority, dca, dca_info, waterless_dict)
                    if case_eval['smart']:
                        single_case_factors = build_single_case_factors(b, dca, factors, 'mp')
                        areas = {x: dca_info.loc[dca]['area_sqmi'] * single_case_factors[x] \
                                for x in guild_list}
                        change = (case_eval['benefit1'], case_eval['benefit2'], b, \
                                dca, flag, areas)
                        smart_cases.append(change)
        smart_cases = sorted(smart_cases, key=lambda x: (x[0], x[1]), \
                reverse=True)
        try:
            hab_checks = guild_list
            while True:
                possible_changes = len(smart_cases)
                best_change = smart_cases[0]
                test_case = case.copy()
                test_case.loc[best_change[3]] = np.array(best_change[2])
                case_factors = build_case_factors(test_case, factors, 'mp')
                test_total = case_factors.multiply(dca_info['area_ac'], axis=0).sum()
                test_percent = test_total/total['base']
                other_dca_smart_cases = [x for x in smart_cases if x[3] != best_change[3]]
                guild_available = get_guild_available(other_dca_smart_cases)
                violate_flag = check_exceed_area(test_total, new_total, hab_limits, guild_available)
                bc_area = best_change[5]
                bc_old_area = {x: get_area(case.loc[best_change[3]].tolist(), best_change[3], \
                        factors, dca_info, x) for x in guild_list}
                pass_continue = False
                for hab in hab_checks:
                    bc_change = bc_area[hab] - bc_old_area[hab]
                    if violate_flag[hab] == 'over':
                        smart_cases = [x for x in smart_cases if \
                                x[5][hab] - bc_old_area[hab] < bc_change]
                        hab_checks = [x for x in hab_checks if x != hab]
                        pass_continue = True
                    if violate_flag[hab] == 'under' and bc_change <= 0:
                        smart_cases = [x for x in smart_cases if \
                                x[5][hab] - bc_old_area[hab] > 0]
                        output = "eliminating " + \
                                str(possible_changes - len(smart_cases)) + " of " + \
                                str(possible_changes) + " possible changes." + " (" + \
                                hab + " pushed " + violate_flag[hab] + \
                                " target area range)"
                        print output
                        log_file.write(output + "\n")
                        pass_continue = True
                if pass_continue:
                    continue
                if best_change[4] == 'soft':
                    if soft_transition + \
                        dca_info.loc[best_change[3]]['area_sqmi'] > soft_limit:
                        smart_cases = [x for x in smart_cases if \
                                x[4] != 'soft' or \
                                dca_info.loc[x[3]]['area_sqmi'] < \
                                dca_info.loc[best_change[3]]['area_sqmi']]
                        output = "eliminating " + str(possible_changes - len(smart_cases)) + \
                                " of " + str(possible_changes) + " possible changes." + \
                                " (soft transition limit exceeded)"
                        print output
                        log_file.write(output + "\n")
                        continue
                    else:
                        soft_transition += dca_info.loc[best_change[3]]['area_sqmi']
#                        if dcm_type in dcm_limits.keys():
#                            dcm_area_tracking[dcm_type] += \
#                                    dca_info.loc[best_change[3]]['area_sqmi']
                        break
                else:
                    if hard_transition + \
                        dca_info.loc[best_change[3]]['area_sqmi'] > hard_limit:
                        smart_cases = [x for x in smart_cases if \
                                x[4] != 'hard' or \
                                dca_info.loc[x[3]]['area_sqmi'] < \
                                dca_info.loc[best_change[3]]['area_sqmi']]
                        output = "eliminating " + str(possible_changes - len(smart_cases)) + \
                                " of " + str(possible_changes) + " possible changes." + \
                                " (hard transition limit exceeded)"
                        print output
                        log_file.write(output + "\n")
                        continue
                    else:
                        hard_transition += dca_info.loc[best_change[3]]['area_sqmi']
#                        if dcm_type in dcm_limits.keys():
#                            dcm_area_tracking[dcm_type] += \
#                                    dca_info.loc[best_change[3]]['area_sqmi']
                        break
        except:
            break
        prior_assignment = case.loc[best_change[3]].tolist()
        prior_dcm = case.columns.tolist()[prior_assignment.index(1)]
        case.loc[best_change[3]] = np.array(best_change[2])
        constraints.loc[best_change[3]] = np.array(best_change[2])
        assignments = get_assignments(case, dca_list, dcm_list)
        assignments.columns = ["step"+str(step)]
        new_total = case_factors.multiply(dca_info['area_ac'], axis=0).sum()
        new_percent = new_total/total['base']
        priority = prioritize(new_percent, hab_limits)
        target_flag = check_exceed_area(test_total, new_total, hab_limits, guild_available)
        log_file.write(best_change[3] + " from " + prior_dcm + " to " + \
            case.columns.tolist()[best_change[2].index(1)] + "\n")
        tracking = tracking.append({'dca': best_change[3], \
                'mp': case.columns.tolist()[best_change[2].index(1)], \
                'step': step}, ignore_index=True)
    step_info[step] = {'totals': new_total, 'percent_base': new_percent, \
            'hard': hard_transition, \
            'soft': soft_transition, \
            'assignments': assignments}
total_water_savings = total['step0']['water'] - new_total['water']
print 'Finished!'
print 'Total Water Savings = ' + str(total_water_savings) + ' acre-feet/year'

tracking = tracking.set_index('dca', drop=True)
assignment_output = step_info[0]['assignments']
for i in range(1, 6):
    assignment_output["step"+str(i)] = step_info[i]['assignments']
assignment_output = assignment_output.join(tracking)
assignment_output['mp'] = [x if str(y) == 'nan' else y for x, y in \
        zip(assignment_output['step5'], assignment_output['mp'])]
assignment_output['step'] = [0 if str(x) == 'nan' else x for x in \
        assignment_output['step']]
assignment_output.to_csv(output_csv)

hab2dcm = mp_file.parse(sheet_name="Cost Analysis Input", header=0, \
        usecols="I,J,K,L").dropna(how='any')
hab2dcm = hab2dcm.append({'mp_name':'total', 'desc':'x', 'hab_id':'x', 'dust_dcm':'x'},
        ignore_index=True)
dcm_order = mp_file.parse(sheet_name="Cost Analysis Input", header=0, \
        usecols="E")[:11]['dust_dcm'].tolist()
dcm_order.append('total')
hab_dict = pd.Series(hab2dcm.dust_dcm.values, index=hab2dcm.mp_name)
summary_df = assignment_output.join(dca_info[['area_sqmi', 'area_ac']])
summary_melt = pd.melt(summary_df, id_vars=['area_ac'], \
        value_vars=['step'+str(i) for i in range(0, 6)], \
        var_name='step', value_name='mp_name')
summary_melt['dcm'] = summary_melt['mp_name']
summary_melt['dcm'].replace(hab_dict, inplace=True)
summary = {'mp_name': [], 'dcm': []}
for nm in summary.keys():
    summary[nm] = summary_melt.groupby([nm, 'step'])['area_ac'].agg('sum').unstack()
    tot = summary[nm].sum().rename('total')
    summary[nm] = summary[nm].append(tot)
    summary[nm].fillna(0, inplace=True)
summary['dcm'] = summary['dcm'].reindex(dcm_order).copy().drop('None')
hab2dcm.set_index('mp_name', inplace=True)
summary['mp_name'] = summary['mp_name'].join(hab2dcm, how='right')
summary['mp_name'].drop(['desc', 'hab_id', 'dust_dcm'], axis=1, inplace=True)
summary['mp_name'].fillna(0, inplace=True)
summary['mp_name'] = summary['mp_name'].drop('None')

# write results into output workbook
wb = load_workbook(filename = file_path + file_name)
ws = wb['MP_new']
for i in range(0, len(assignment_output), 1):
    offset = 22
    ws.cell(row=i+offset, column=7).value = assignment_output['mp'][i]
    ws.cell(row=i+offset, column=8).value = assignment_output['step'][i]
rw = 3
col_ind = {'bw':2, 'mw':3, 'pl':4, 'ms':5, 'md':6, 'water':7}
for j in ['base', 0, 1, 2, 3, 4, 5]:
    for k in col_ind.keys():
        ws.cell(row=rw, column=col_ind[k]).value = step_info[j]['totals'][k]
    rw += 1
# write area summary tables
ws = wb['Area Summary']
for i in range(0, len(summary['dcm']), 1):
    for j in range(0, 6):
        ws.cell(row=i+5, column=j+2).value = int(summary['dcm'].iloc[i, j].round())
for i in range(0, len(summary['mp_name']), 1):
    for j in range(0, 6):
        ws.cell(row=i+5, column=j+10).value = int(summary['mp_name'].iloc[i, j].round())
wb.save(output_excel)
book = load_workbook(filename=output_excel)
writer = pd.ExcelWriter(output_excel, engine = 'openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

writer.save()

