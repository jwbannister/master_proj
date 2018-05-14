#!/usr/bin/env python

import pandas as pd
import numpy as np
import datetime
from itertools import product
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.utils import range_boundaries

def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
patch_worksheet()

# path for excel files
file_path = "/home/john/code/master_proj/"
npv_name = "MP NPV TEMPLATE.xlsx"
mp_name = "output/MP Workbook 05_14_18 09_22.xlsx"

# read reference tables from NPY Calculation workbook
input_file = pd.ExcelFile(file_path + npv_name)
# what years are MP steps to be implemented?
step_years = input_file.parse(sheet_name="Script Input", header=0, \
        usecols="A,B", converters={'year':int}).dropna(how='any')
step_dict = pd.Series(step_years.year.values, index=step_years.step)
# list of years for projection
year_range = range(int(step_years.year.min()), int(step_years.year.max()) + 21, 1)
# DCM costs
dcm_costs = input_file.parse(sheet_name="Script Input", header=0, \
        usecols="D,E,F").dropna(how='any')
dcm_costs.dropna(inplace=True)
dcm_costs.set_index('dust_dcm', inplace=True)
# mapping between MP habitats and DCMs
hab2dcm = input_file.parse(sheet_name="Script Input", header=0, \
        usecols="H,I,J,K").dropna(how='any')
hab_dict = pd.Series(hab2dcm.dust_dcm.values, index=hab2dcm.mp_name)

# read data from John Dickey's workbook
mp_file = pd.ExcelFile(file_path + mp_name)
# MP implementation schedule
mp_new_names = ["dca", "acres", "base", "dwm", "step0", "step5", "step"]
mp_new = mp_file.parse(sheet_name="MP_new", header=20, \
        usecols="A,B,D,E,F,G,H", names=mp_new_names, \
        converters={'Step':int}).dropna(how='any')
base_water = mp_file.parse(sheet_name="MP_new", header=0, \
        usecols="G").iloc[0, 0]
# pull water demand by DCA (in acre-ft/year)
mp_steps_wd = mp_file.parse(sheet_name="Script Output - Step WD", header=0, \
        usecols="A,B,C,D,E,F,G", \
        names=['dca', 'step0', 'step1', 'step2', 'step3', 'step4', 'step5'])
mp_steps_wd.set_index('dca', inplace=True)

# get dca areas
dca_areas = mp_new[['dca', 'acres']].copy().set_index('dca')
dca_areas['miles'] = dca_areas['acres'] * 0.0015625

# expand data to show all steps in Master Project
mp_steps = mp_new.copy()
mp_steps.set_index('dca', inplace=True)
for n in ['1', '2', '3', '4']:
    mp_steps['step' + n] = ['X'] * len(mp_steps)
    for i in mp_steps.index:
        if mp_steps.loc[i, 'step'] > int(n):
            mp_steps.loc[i, 'step' + n] = mp_steps.loc[i, 'step0']
        else:
            mp_steps.loc[i, 'step' + n] = mp_steps.loc[i, 'step5']
mp_steps.drop(['acres', 'base', 'dwm', 'step'], axis=1, inplace=True)

mp_years = mp_steps.copy()
mp_years_wd = mp_steps_wd.copy()
for obj2 in [mp_years, mp_years_wd]:
    obj2.columns = [step_dict[a] for a in obj2.columns]
    # expand table to include all years in range
    for yr in year_range:
        if yr in obj2.columns:
            obj2[yr] = obj2[yr]
        else:
            obj2[yr] = obj2[yr-1]

mp_years_dcm = mp_years[year_range].copy()
mp_years_dcm.replace(hab_dict, inplace=True)
# build costs array 
# dim 0 = total capital cost ($ million) 
# dim 1 = o&m cost ($ million) 
mp_costs = {'capital': mp_years_dcm.copy(), 'om': mp_years_dcm.copy()}
mp_costs['om'].loc[:, 2018] = [dcm_costs.loc[x]['om'] \
        for x in mp_years_dcm[2018].tolist()]
mp_costs['capital'].loc[:, 2018] = [0 for x in mp_years_dcm[2018].tolist()]
for i in mp_years_dcm.index.tolist():
    for j in year_range[1:]:
        if mp_years_dcm.loc[i, j-1] == mp_years_dcm.loc[i, j]:
            mp_costs['om'].loc[i, j] = dcm_costs.loc[mp_years_dcm.loc[i, j]]['om']
            mp_costs['capital'].loc[i, j] = 0
        else:
            mp_costs['capital'].loc[i, j] = \
                    dcm_costs.loc[mp_years_dcm.loc[i, j]]['capital']
            mp_costs['om'].loc[i, j] = 0

cost_summary = pd.DataFrame({\
        'capital': mp_costs['capital'].sum(axis=0), \
        'om': mp_costs['om'].sum(axis=0), \
        'water_demand': mp_years_wd.sum(axis=0)
        })
cost_summary['year'] = cost_summary.index.tolist()

sheet_dict = {'step0':'Step0', 'step1':'Step1', 'step2':'Step2', 'step3':'Step3', \
        'step4':'Step4', 'step5':'Full Project'}
wb = load_workbook(filename = file_path + npv_name)
for step in step_years.step:
    yr = int(step_years[step_years.step==step].year.item())
    ind = cost_summary.year.tolist().index(yr)
    capital_output = cost_summary.capital.tolist()
    capital_output[ind:] = [0]*(len(capital_output) - ind)
    om_output = cost_summary.om.tolist()
    om_output[ind:] = [om_output[ind]]*(len(om_output) - ind)
    wd_output = cost_summary.water_demand.tolist()
    wd_output[ind:] = [wd_output[ind]]*(len(wd_output) - ind)
    ws = wb[sheet_dict[step]]
    for i in range(0, len(cost_summary), 1):
        offset = 12
        ws.cell(row=i+offset, column=3).value = capital_output[i]
        ws.cell(row=i+offset, column=4).value = om_output[i]
        ws.cell(row=i+offset, column=19).value = wd_output[i]

    mp_steps_dcm = pd.concat([mp_steps.replace(hab_dict), \
            dca_areas['acres']], axis=1)
    mp_steps_hab = pd.concat([mp_steps, dca_areas['acres']], axis=1)
    # write water demand summary tables
    ws = wb['Water Use Summary']
    wd_summary_dcm = pd.DataFrame({'wd':mp_steps_wd[step], \
            'dcm':mp_steps_dcm[step]}).groupby('dcm')['wd'].sum()
    wd_summary_dcm = wd_summary_dcm.reindex(dcm_costs.index.tolist())
    wd_summary_dcm.fillna(0, inplace=True)
    wd_summary_dcm = \
            wd_summary_dcm.append(pd.Series({'total':wd_summary_dcm.sum()}))
    for i in range(0, len(wd_summary_dcm), 1):
        col = step_years.index[step_years.step==step].item()
        ws.cell(row=i+5, column=col+2).value = int(wd_summary_dcm[i].round())
    wd_summary_hab = pd.DataFrame({'wd':mp_steps_wd[step], \
            'hab':mp_steps[step]}).groupby('hab')['wd'].sum()
    wd_summary_hab = wd_summary_hab.reindex(hab2dcm.mp_name.tolist())
    wd_summary_hab.fillna(0, inplace=True)
    wd_summary_hab = \
            wd_summary_hab.append(pd.Series({'total':wd_summary_hab.sum()}))
    for i in range(0, len(wd_summary_hab), 1):
        col = step_years.index[step_years.step==step].item()
        ws.cell(row=i+5, column=col+10).value = int(wd_summary_hab[i].round())

    # write area summary tables
    ws = wb['Area Summary']
    area_summary_dcm = pd.DataFrame({'acres':mp_steps_dcm['acres'], \
            'dcm':mp_steps_dcm[step]}).groupby('dcm')['acres'].sum()
    area_summary_dcm = area_summary_dcm.reindex(dcm_costs.index.tolist())
    area_summary_dcm.fillna(0, inplace=True)
    area_summary_dcm = \
            area_summary_dcm.append(pd.Series({'total':area_summary_dcm.sum()}))
    for i in range(0, len(area_summary_dcm), 1):
        col = step_years.index[step_years.step==step].item()
        ws.cell(row=i+5, column=col+2).value = int(area_summary_dcm[i].round())
    area_summary_hab = pd.DataFrame({'acres':mp_steps_dcm['acres'], \
            'hab':mp_steps_hab[step]}).groupby('hab')['acres'].sum()
    area_summary_hab = area_summary_hab.reindex(hab2dcm.mp_name.tolist())
    area_summary_hab.fillna(0, inplace=True)
    area_summary_hab = \
            area_summary_hab.append(pd.Series({'total':area_summary_hab.sum()}))
    for i in range(0, len(area_summary_hab), 1):
        col = step_years.index[step_years.step==step].item()
        ws.cell(row=i+5, column=col+12).value = int(area_summary_hab[i].round())

ws = wb['NPV Summary']
ws.cell(row=16, column=2).value = base_water
ws.cell(row=16, column=3).value = "Base Case water usage (acre-feet/year)"
ws.cell(row=18, column=2).value = "Data read from Master Project workbook '" + \
        mp_name + "'"
ws.cell(row=19, column=2).value = 'NPV Analysis run on ' + \
        datetime.datetime.now().strftime('%m-%d-%Y %H:%M')
output_file = file_path + "output/" + npv_name[:7] + \
        datetime.datetime.now().strftime('%m_%d_%y %H_%M') + '.xlsx'
wb.save(output_file)

book = load_workbook(filename=output_file)
writer = pd.ExcelWriter(output_file, engine = 'openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
mp_new.to_excel(writer, sheet_name='MP Schedule', index=False)
mp_steps_wd.to_excel(writer, sheet_name='MP Water', index=False)
writer.save()
