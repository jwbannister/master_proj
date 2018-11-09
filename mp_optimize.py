#! /usr/bin/env python
import pandas as pd
import numpy as np
import datetime
import os
import sys
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/364/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A0:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/364/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A0:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+0)
        # cols = range(min_col, max_col+0)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 0, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet.merge_cells = merge_cells
patch_worksheet()

def build_factor_tables(water_adjust=1):
    design = mp_file.parse(sheet_name="Design HV & WD", header=2, \
            usecols="A,D,E,F,G,H,I", \
            names=["dcm", "bw", "mw", "pl", "ms", "md", "water"])[:len(dcm_list)]
    design.dropna(how='all', inplace=True)
    design.set_index('dcm', inplace=True)
    design['water'] = design['water'] * water_adjust
    # build up custom habitat and water factor tables
    asbuilt = mp_file.parse(sheet_name="As-Built HV & WD", header=0, \
            usecols="A,B,C,D,E,F,G,H", \
            names=["dca", "dcm", "bw", "mw", "pl", "ms", "md", "water"])
    asbuilt.dropna(how='all', inplace=True)
    asbuilt_water_ind = [(dca, dcm) for dca, dcm, water in \
            zip(asbuilt['dca'], asbuilt['dcm'], asbuilt['water']) \
            if water != ' - ']
    asbuilt['water'] = [design.loc[d]['water'] if w == ' - ' \
            else w for d, w in zip(asbuilt['dcm'], asbuilt['water'])]
    asbuilt.set_index(['dca', 'dcm'], inplace=True)
    factors = {'asbuilt':asbuilt.copy(), 'design':design.copy()}
    return factors, asbuilt_water_ind

def read_dca_info():
    dca_info = mp_file.parse(sheet_name="MP_new", header=None, skiprows=21, \
            usecols="A,B,C,D,E,G", \
            names=["dca", "area_ac", "area_sqmi", "phase", "base", "step0"])
    dca_info.set_index('dca', inplace=True)
    return dca_info

def build_past_status(stp):
    state = dca_info[stp]
    stp_factors = get_factors(state)
    stp_status = stp_factors.join(dca_info[['area_ac', 'area_sqmi']], on='dca')
    cols = [x + "_ac" if x != 'water' else x + "_af/y" for x in stp_factors.columns]
    for i in stp_factors.columns:
        col_name = i + "_ac" if i != 'water' else i + "_af/y"
        stp_status[col_name] = stp_factors[i] * stp_status['area_ac']
    return stp_status

def get_factors(state):
    a = zip(state.index, state)
    b = pd.DataFrame()
    for i in a:
        try:
            c = factors['asbuilt'].loc[i[0], i[1]]
        except:
            c = factors['design'].loc[i[1]]
        d = pd.DataFrame(c).transpose()
        d['dca'] = i[0]
        d['dcm'] = i[1]
        b = b.append(d)
    b.set_index(['dca', 'dcm'], inplace=True)
    return b

def define_soft():
    soft_dcm_input = mp_file.parse(sheet_name="MP Analysis Input", header=6, \
            usecols="A").iloc[:, 0].tolist()[4:11]
    soft_dcms = [x for x in soft_dcm_input if x in dcm_list]
    soft_idx = [x for x, y in enumerate(dcm_list) if y in soft_dcms]
    return soft_idx

def evaluate_dca_change(dca, dcm, state, factors, \
        priority, waterless_dict):
    current_factors = state.loc[dca][factor_keys].squeeze()
    new_factors = factors['design'].loc[dcm]
    if priority[1]=='water':
        smart = new_factors['water'] - current_factors['water'] < 0
        benefit1 = current_factors['water'] - new_factors['water']
        try:
            benefit2 = waterless_dict[dcm]
        except:
            benefit2 = -6
    else:
        p1_increase = new_factors[priority[1]] - current_factors[priority[1]]
        water_increase = 100 + (new_factors['water'] - current_factors['water'])
        smart = p1_increase > 0
        benefit1 = p1_increase / water_increase
        p2_increase = new_factors[priority[2]] - current_factors[priority[2]]
        benefit2 = p2_increase / water_increase
    return {'smart': smart, 'benefit1':benefit1, 'benefit2':benefit2}

def prioritize(value_percents, hab_minimums):
    hab_deficits = {x: value_percents[x] - hab_minimums[x] \
            for x in value_percents.index.tolist() \
            if x != 'water'}
    if any([x < 0 for x in hab_deficits.values()]):
        sort_deficits = sorted(hab_deficits.values())
        one = hab_deficits.values().index(sort_deficits[0])
        two = hab_deficits.values().index(sort_deficits[1])
        return {1: hab_deficits.keys()[one], 2: hab_deficits.keys()[two]}
    else:
        return {1: 'water', 2: value_percents[0:5].idxmin()}

def try_get_dcm(row):
    try:
        return dcm_list[row.tolist().index(1)]
    except:
        return dca_info.loc[row.name, 'step0']

def get_assignments(case, dca_list, dcm_list):
    assignments = pd.DataFrame([try_get_dcm(row) \
            for index, row in case.iterrows()], index=dca_list, columns=['dcm'])
    return assignments

def printout(flag):
    readout = ""
    if flag == 'screen':
        for x in new_percent.keys().tolist():
            if priority[1] == x:
                pri = color.BOLD
            else:
                pri = ""
            try:
                if new_percent[x] >= hab_limits[x]:
                    readout = readout + pri + color.GREEN + x + ": " + \
                            str(round(new_percent[x], 3)) \
                            + ", " + color.END
                else:
                    readout = readout + pri + color.RED + x + ": " + \
                            str(round(new_percent[x], 3)) \
                            + ", " + color.END
            except:
                readout = readout + pri + x + ": " + \
                        str(round(new_percent[x], 3)) \
                        + ", " + color.END
        return readout
    else:
        for x in new_percent.keys().tolist():
            readout = readout + x + ": " + \
                    str(round(new_percent[x], 3)) \
                    + ", "
        readout = readout + "priority1 = " + priority[1] + ", priority2 = " +\
                priority[2]
    return readout

def check_exceed_area(test_totals, new_totals, limits, guild_available, \
        meadow_limits=True, bw_limit=True):
    exceed_flag = {x: 'ok' for x in guild_list if x != 'water'}
    for x in exceed_flag.keys():
        if (test_totals[x] + (guild_available[x] * 640)) / total['base'][x] < limits[x]:
            exceed_flag[x] = 'under'
    # meadow is hard to establish, do not want to reduce only to have to
    # re-establish. Prevent meadow from dipping below target value and never
    # add any more meadow
    if not unconstrained_case and meadow_limits:
        if test_totals['md'] / total['base']['md'] < limits['md']:
            exceed_flag['md'] = 'under'
        if test_totals['md'] > new_totals['md'] :
            exceed_flag['md'] = 'over'
    # no design habitats to create more breeding waterfowl areas, so don't allow
    # breeding waterfowl to fall below target
    if not unconstrained_case and bw_limit:
        if test_totals['bw'] / total['base']['bw'] < limits['bw']:
            exceed_flag['bw'] = 'under'
    return exceed_flag

def set_constraint(axis, idx, new_constraint, constraint_df):
    new = new_constraint
    if axis == 1:
        existing = constraint_df.loc[idx].tolist()
        constraint_df.loc[idx] = [min(x, y) for x, y in zip(new, existing)]
    else:
        existing = constraint_df[idx]
        constraint_df[idx] = [min(x, y) for x, y in zip(new, existing)]
    return constraint_df

def get_guild_available(best_change, approach_factor=0.9):
    dca_considered = best_change[3]
    if best_change[4] == 'hard':
        new_hard = dca_info.loc[best_change[3]]['area_sqmi']
    else:
        new_hard = 0
    possible_cases = generate_possible_changes(smart_only=False)
    possible_cases = [x for x in possible_cases if x[3] != dca_considered]
    guild_available = {}
    available_hard = trans_limits['hard'] - trans_area['hard'] - new_hard
    for hab in guild_list:
        temp = []
        for dca in set([x[3] for x in possible_cases]):
            try:
                temp.append(\
                        sorted([[x[6][hab], dca_info.loc[dca]['area_sqmi']] \
                        for x in possible_cases \
                        if x[3] == dca and \
                        dca_info.loc[dca]['area_sqmi'] < available_hard], \
                        reverse=True)[0])
            except:
                temp.append([0, 0])
        temp = [x for x in sorted(temp, reverse=True) if x[1] < available_hard]
        temp = [x for x in temp if x[0] > 0]
        temp1 = sorted([[x[0]/x[1], x[0], x[1]] for x in temp], reverse=True)
        while sum([x[2] for x in temp1]) > approach_factor * available_hard:
            temp1.pop()
        guild_available[hab] = sum([x[1] for x in temp1])
    return guild_available

def initialize_constraints():
    dcm_constraints = pd.DataFrame(np.ones((len(dca_list), len(dcm_list)), \
            dtype=np.int8))
    dcm_constraints.index = dca_list
    dcm_constraints.columns = dcm_list
    step_constraints = pd.DataFrame(np.ones((len(dca_list), 5), dtype=np.int8))
    step_constraints.index = dca_list
    step_constraints.columns = range(1, 6)
    return dcm_constraints, step_constraints

def update_constraints(dcm_constraints, step_constraints):
    # set hard-wired constraints
    # Constraint to only remove "Meadow" habitat and to limit BW re-build is 
    # wired into check_exceed_area function
    # nothing allowed as Enhanced Natural Vegetation (ENV) except Channel Areas
    new = [1 if 'Channel' in x else 0 for x in dca_list]
    dcm_constraints = set_constraint(0, 'ENV', new, dcm_constraints)
    # freeze Channel areas
    new = [1 if x=='ENV' else 0 for x in dcm_list]
    for dca in [x for x in dca_list if 'Channel' in x]:
        dcm_constraints = set_constraint(1, dca, new, dcm_constraints)
    # no additional sand fences besides existing T1A1
    new = [1 if x == 'T1A-1' else 0 for x in dca_list]
    dcm_constraints = set_constraint(0, 'Sand Fences', new, dcm_constraints)
    # all DCAs currently under waterless DCM should remain unchanged
    waterless = dca_info.loc[[x in waterless_dict.keys() + \
            ['Brine (DWM)', 'Sand Fences'] \
            for x in dca_info['step0']]]
    for dca in waterless.index:
        new = [1 if x == waterless.loc[dca]['step0'] else 0 for x in dcm_list]
        dcm_constraints = set_constraint(1, dca, new, dcm_constraints)
    # all DCAs under dust control, no new "None" DCMs allowed
    new = [0 for x in dca_list]
    dcm_constraints = set_constraint(0, 'None', new, dcm_constraints)
    # no new DWM areas allowed
    dwm_list = [x for x in dcm_list if 'DWM' in x]
    dwm_dcas = [x for x, y in zip(dca_info.index.tolist(), \
            dca_info['step0']) if 'DWM' in y]
    for dcm in dwm_list:
        new = [1 if x in dwm_dcas else 0 for x in dca_list]
        dcm_constraints = set_constraint(0, dcm, new, dcm_constraints)

    # read in and implement Ops constraints from LAUNCHPAD
    constraints_input = mp_file.parse(sheet_name="Constraints Input", header=7, \
            usecols="A:J")
    constraints_input.set_index('dca', inplace=True)
    step_start = ['Step' in str(x) for \
            x in constraints_input.index.tolist()].index(True)
    phase_start = ['Phase' in str(x) for \
            x in constraints_input.index.tolist()].index(True)
    dcm_constraint_input = constraints_input.iloc[:step_start, :].dropna(how='all')
    step_constraint_input = constraints_input.iloc[step_start+1:phase_start, \
            :].dropna(how='all')
    phase_constraint_input = constraints_input.iloc[phase_start+1:, :].dropna(how='all')
    phase_constraint_input.index = [str(x) for x in phase_constraint_input.index]
    for dca in dcm_constraint_input.index:
        if dcm_constraint_input.loc[dca, 'type']=='not':
            not_allowed = dcm_constraint_input.loc[dca].dropna().tolist()[1:]
            new = [0 if x in not_allowed else 1 for x in dcm_list]
        else:
            allowed = dcm_constraint_input.loc[dca].dropna().tolist()[1:]
            new = [1 if x in allowed else 0 for x in dcm_list]
        dcm_constraints = set_constraint(1, dca, new, dcm_constraints)
    for dca in step_constraint_input.index:
        if step_constraint_input.loc[dca, 'type']=='not':
            not_allowed = step_constraint_input.loc[dca].dropna().tolist()[1:]
            new = [0 if x in not_allowed else 1 for x in range(1, 6)]
        else:
            allowed = step_constraint_input.loc[dca].dropna().tolist()[1:]
            new = [1 if x in allowed else 0 for x in range(1, 6)]
        step_constraints = set_constraint(1, dca, new, step_constraints)
    for phase in phase_constraint_input.index.tolist():
        if phase_constraint_input.loc[phase, 'type']=='not':
            not_allowed = phase_constraint_input.loc[phase].dropna().tolist()[1:]
            new = [0 if x in not_allowed else 1 for x in range(1, 6)]
        else:
            allowed = phase_constraint_input.loc[phase].dropna().tolist()[1:]
            new = [1 if x in allowed else 0 for x in range(1, 6)]
        phase_dcas = dca_info.loc[dca_info['phase'] == float(phase)].index
        for dca in phase_dcas:
            step_constraints = set_constraint(1, dca, new, step_constraints)
    return dcm_constraints, step_constraints

def generate_possible_changes(smart_only=True, force_change=False):
    smart_cases = []
    for dca in dca_list:
        if force_change or step_constraints.loc[dca, step] != 0:
            if force_change:
                tmp = [1 if 'asbuilt' not in x and '(DWM)' not in x and \
                        '(improved)' not in x else 0 \
                        for x in dcm_list ]
            else:
                tmp = constraints.loc[dca].tolist()
            tmp_ind = [x for x, y in enumerate(tmp) if y == 1]
            for dcm_ind in tmp_ind:
                if dcm_ind in soft_idx:
                    flag = 'soft'
                else:
                    flag='hard'
                b = [0 for x in tmp]
                b[dcm_ind] = 1
                dcm = dcm_list[dcm_ind]
                case_eval = evaluate_dca_change(dca, dcm, new_state, factors, priority, \
                        waterless_dict)
                if smart_only and not case_eval['smart']:
                    continue
                else:
                    new_factors = factors['design'].loc[dcm]
                    new_areas = {x: dca_info.loc[dca]['area_sqmi'] * new_factors[x] \
                            for x in factor_keys}
                    current_factors = new_state.loc[dca][factor_keys].squeeze()
                    current_areas = {x: dca_info.loc[dca]['area_sqmi'] * \
                            current_factors[x] for x in factor_keys}
                    guild_changes = {x: new_areas[x] - current_areas[x] \
                            for x in factor_keys}
                    change = (case_eval['benefit1'], case_eval['benefit2'], b, \
                            dca, flag, current_areas, guild_changes, dcm)
                    smart_cases.append(change)
    smart_cases = sorted(smart_cases, key=lambda x: (x[0], x[1]), \
            reverse=True)
    return smart_cases

def check_guild_violations(smart_cases, best_change, meadow_limits=True):
    violate_flag = check_exceed_area(test_total, new_total, hab_limits, \
            guild_available, meadow_limits=meadow_limits)
    violations = [x for x in guild_list \
            if (violate_flag[x] == 'under' and best_change[6][x] < 0) \
            or (violate_flag[x] == 'over' and best_change[6][x] > 0)]
    if len(violations) == 0: return smart_cases, 'NA', 'NA'
    hab = violations[0]
    comp = {'over': lambda x,y: x < y, 'under':lambda x,y: x > y}
    filtered_cases = [x for x in smart_cases if \
            comp[violate_flag[hab]](x[6][hab], best_change[6][hab])]
    return filtered_cases, hab, violate_flag[hab]

def initialize_files():
    file_path = os.path.realpath(os.getcwd()) + "/"
    file_name = "MP LAUNCHPAD.xlsx"
    mp_file = pd.ExcelFile(file_path + file_name)
    timestamp = datetime.datetime.now().strftime('%m_%d_%y %H_%M')
    output_excel = file_path + "output/" + "MP " + timestamp + file_flag + '.xlsx'
    output_csv = file_path + "output/mp_steps " + timestamp + file_flag + '.csv'
    # read in current state of workbook for future writing
    wb = load_workbook(filename = file_path + file_name)
    return mp_file, output_excel, output_csv, wb

# set algorithm options and filename flags
efficient_steps = True #stop step if water savings plateaus
unconstrained_case = False #remove all constraints
freeze_farm = True #keep "farm" managed veg as is
factor_water = True #adjust water useage values so base water matches preset value
preset_base_water = 73351
truncate_steps = True #erase step changes if no water savings is acheived
force = False #force changes
file_flag = ""
if unconstrained_case: file_flag = file_flag + " NO_CONSTRAINTS"
if not efficient_steps: file_flag = file_flag + " EFFICIENT_STEPS_OFF"
if not freeze_farm: file_flag = file_flag + " FARM_FROZEN_OFF"
if not factor_water: file_flag = file_flag + " H20_ADJUST_OFF"
if force: file_flag = file_flag + " FORCED_CHANGES"

# read data from original Master Project planning workbook
mp_file, output_excel, output_csv, wb = initialize_files()

design_dcms = mp_file.parse(sheet_name="Design HV & WD", header=2, \
        usecols="A,C").dropna(how='any')
dcm_dict = pd.Series(design_dcms.Type.values, index=design_dcms.MP_id)

dca_info = read_dca_info()
dca_list = dca_info.index.tolist()
dcm_list = [x for x in design_dcms['MP_id'] if not any([y in x for y in \
        ['(DWM)', 'improved', 'as-built']])]
hab_list = [design_dcms['MP_id'][idx] for idx, i \
        in enumerate(design_dcms['Type'][:len(dcm_list)]) if i=='Habitat DCM']
factors, asbuilt_water_ind = build_factor_tables()
guild_list = [x for x in factors['design'].columns if x != 'water']
factor_keys = [x for x in factors['design'].columns]

# read and set preferences for waterless DCMs
waterless_preference = mp_file.parse(sheet_name="MP Analysis Input", header=18, \
        usecols="A").iloc[:, 0].tolist()
waterless_dict = {x:-y for x, y in zip(waterless_preference, range(1, 6))}

# generate list of "soft transition" DCMS
soft_idx = define_soft()

# read limits and toggles
trans_limit_input = mp_file.parse(sheet_name="MP Analysis Input", header=None, \
        usecols="B").iloc[:, 0].tolist()[0:4]
trans_limits = {'hard': trans_limit_input[0], 'soft': trans_limit_input[1]}
hab_limit_input = mp_file.parse(sheet_name="MP Analysis Input", \
        usecols="B,C").iloc[5:10, 0:2]
hab_limits = dict(zip(hab_limit_input.iloc[:, 1].tolist(), \
        hab_limit_input.iloc[:, 0].tolist()))

dcm_constraints, step_constraints = initialize_constraints()

if not unconstrained_case:
    dcm_constraints, step_constraints = \
            update_constraints(dcm_constraints, step_constraints)

if freeze_farm:
    farm_dcas = dca_info.loc[[x == 'Veg 08' for x in dca_info['step0']]]
    new = [0 for x in range(1, 6)]
    for dca in farm_dcas.index:
        step_constraints = set_constraint(1, dca, new, step_constraints)

if force:
    forces = mp_file.parse(sheet_name="MP Analysis Input", header=0, skiprows=1, \
            usecols="J,K,L")
    forces.dropna(how='any', inplace=True)
    forces.set_index('dca', inplace=True)
    # prevent forced DCAs from changing before they are forced
    for dca in forces.index:
        new = [1 if x == forces.loc[dca]['step'] else 0 for x in range(1, 6)]
        step_constraints = set_constraint(1, dca, new, step_constraints)

lake_state_pre_water_factor = {x:build_past_status(x) for x in ['base', 'step0']}
if factor_water:
    calc_base_water = lake_state_pre_water_factor['base']['water_af/y'].sum()
    base_asbuilt = [x for x in asbuilt_water_ind \
            if x in lake_state_pre_water_factor['base'].index.get_values().tolist()]
    base_design = [x for x in \
            lake_state_pre_water_factor['base'].index.get_values().tolist() \
            if not (x in asbuilt_water_ind)]
    asbuilt_water_dcas = lake_state_pre_water_factor['base'].loc[base_asbuilt, :]
    asbuilt_water_calc = asbuilt_water_dcas['water_af/y'].sum()
    design_water_dcas = lake_state_pre_water_factor['base'].loc[base_design, :]
    design_water_calc = design_water_dcas['water_af/y'].sum()
    water_adjust = (preset_base_water - asbuilt_water_calc)/design_water_calc
    factors, garbage = build_factor_tables(water_adjust)
lake_state = {x:build_past_status(x) for x in ['base', 'step0']}

total = {}
assignments = {}
for stp in lake_state.keys():
    total[stp] = lake_state[stp][['bw_ac', 'mw_ac', 'pl_ac', 'ms_ac', 'md_ac', \
            'water_af/y']].sum()
    total[stp].index = ['bw', 'mw', 'pl', 'ms', 'md', 'water']

# initialize variables before loop
constraints = dcm_constraints.copy()
new_state = lake_state["step0"].copy()
new_percent = total["step0"]/total['base']
new_total = total["step0"].copy()
priority = prioritize(new_percent, hab_limits)
percent_dict = lambda prcnt: {x: y for x, y in zip(prcnt.keys(), prcnt.values)}
tracking = pd.DataFrame({'dca': 'x', 'from': 'x', 'to': 'x', 'step': 0, \
        'step0_wd': 0, 'mp_wd': 0, 'water_change': 0}, \
        index=[0]).join(\
        pd.DataFrame(percent_dict(new_percent), index=[0]))

recent_water_step = 1.0
change_counter = 1
for step in range(1, 6):
    # intialize step area limits
    trans_area = {x: 0 for x in trans_limits.keys()}
    retry = True
    force_counter = 0
    if force:
        step_forces = forces.loc[forces['step']==step, :]
    while retry:
        if priority[1] == 'water':
            recent_water_step = new_percent['water']
        output = "step " + str(step) + ", change " + str(change_counter) + \
                ": hard/soft " + str(round(trans_area['hard'], 2)) + "/" + \
                str(round(trans_area['soft'], 2))
        print output
        print printout('screen')
        force_trigger = force and force_counter<len(step_forces)
        if force_trigger:
            smart_cases = generate_possible_changes(smart_only=False, force_change=True)
            change_force = step_forces.iloc[force_counter]
            best_change = [x for x in smart_cases if \
                    x[3] == change_force.name and \
                    dcm_list[x[2].index(1)] == change_force['force']][0]
        else:
            smart_cases = generate_possible_changes()
        retry = len(smart_cases) > 0
        while len(smart_cases) > 0:
            if not force_trigger:
                possible_changes = len(smart_cases)
                best_change = smart_cases[0]
                other_dca_smart_cases = [x for x in smart_cases if x[3] != best_change[3]]
            test_case = new_state.copy()
            test_ind = [ind for ind, x in enumerate(test_case.index) if \
                    x[0] == best_change[3]][0]
            tmp_ind = test_case.index.tolist()
            tmp_ind[test_ind] = (best_change[3], best_change[7])
            test_case.index = pd.MultiIndex.from_tuples(tmp_ind, \
                    names=['dca', 'dcm'])
            for i in factor_keys:
                test_case.loc[best_change[3], i] = \
                        factors['design'].loc[best_change[7]][i]
                col_name = i + "_ac" if i != 'water' else i + "_af/y"
                test_case.loc[best_change[3], col_name] = \
                        test_case.loc[best_change[3],'area_ac'].values[0] * \
                        test_case.loc[best_change[3], i].values[0]
            test_total = test_case[['bw_ac', 'mw_ac', 'pl_ac', 'ms_ac', \
                    'md_ac', 'water_af/y']].sum()
            test_total.index = ['bw', 'mw', 'pl', 'ms', 'md', 'water']
            test_percent = test_total/total['base']
            if not force_trigger and priority[1] == 'water' and efficient_steps and \
                    recent_water_step - test_percent['water'] < 0.005:
                smart_cases = [x for x in smart_cases if \
                       x[6]['water'] < best_change[6]['water']]
                output = "eliminating " + str(possible_changes - len(smart_cases)) + \
                        " of " + str(possible_changes) + " possible changes." + \
                        " (inefficient water savings)"
                print output
                retry = len(smart_cases) > 0
                continue
            if not force_trigger and trans_area[best_change[4]] + \
                dca_info.loc[best_change[3]]['area_sqmi'] > trans_limits[best_change[4]]:
                smart_cases = [x for x in smart_cases if \
                        x[4] != best_change[4] or \
                        dca_info.loc[x[3]]['area_sqmi'] < \
                        trans_limits[best_change[4]] - trans_area[best_change[4]]]
                output = "eliminating " + str(possible_changes - len(smart_cases)) + \
                        " of " + str(possible_changes) + " possible changes." + \
                        " (" + best_change[4] + " transition limit exceeded)"
                print output
                retry = len(smart_cases) > 0
                continue
            if not force_trigger:
                guild_available = get_guild_available(best_change, approach_factor=0.9)
                smart_cases, hab, violation = \
                        check_guild_violations(smart_cases, best_change)
            if not force_trigger and len(smart_cases) < possible_changes:
                output = "eliminating " + \
                        str(possible_changes - len(smart_cases)) + " of " + \
                        str(possible_changes) + " possible changes." + " (" + \
                        hab + " pushed " + violation + " target area range)"
                print output
                retry = len(smart_cases) > 0
                continue
            trans_area[best_change[4]] += dca_info.loc[best_change[3]]['area_sqmi']
            constraints.loc[best_change[3]] = np.array(best_change[2])
            new_state = test_case.copy()
            new_total = test_total.copy()
            new_percent = test_percent.copy()
            priority = prioritize(new_percent, hab_limits)
            old_water_duty = lake_state['step0'].loc[best_change[3]]['water'][0]
            new_water_duty = new_state.loc[best_change[3]]['water'][0]
            tracking = tracking.append(pd.DataFrame({'dca': best_change[3], \
                    'from': lake_state['step0'].loc[best_change[3]].index[0], \
                    'to': best_change[7], 'step': step, \
                    'step0_wd_f/y': old_water_duty, 'mp_wd_f/y': new_water_duty, \
                    'water_change_af/y': (new_water_duty - old_water_duty) * \
                    dca_info.loc[best_change[3]]['area_ac']}, \
                    index=[change_counter]).join(\
                       pd.DataFrame(percent_dict(new_percent), index=[change_counter])))
            change_counter += 1
            force_counter += 1
            break
    if total["step" + str(step-1)]['water'] - new_total['water'] < 10 \
            and truncate_steps:
        lake_state["step" + str(step)] = lake_state["step" + str(step-1)]
        total["step" + str(step)] = total["step" + str(step-1)]
        tracking = tracking.loc[tracking['step'] != step]
    else:
        lake_state["step" + str(step)] = new_state.copy()
        total["step" + str(step)] = new_total.copy()
water_min = min([total[x]['water'] for x in total.keys()])
total_water_savings = total['step0']['water'] - water_min
print 'Finished!'
print 'Total Water Savings = ' + str(total_water_savings) + ' acre-feet/year'

# tracking = tracking.set_index('dca', drop=True)
tracking = tracking[['dca','from', 'to', 'step', 'bw', 'mw', 'pl', 'ms', 'md', \
        'water', 'step0_wd_f/y', 'mp_wd_f/y', 'water_change_af/y']]
assignment_output = lake_state['base'][['area_ac', 'area_sqmi']]
assignment_output = assignment_output.join(tracking.set_index('dca').drop('x'))
assignment_output['base'] = assignment_output.index.get_level_values('dcm')
for i in range(0,6):
    stp = "step" + str(i)
    assignment_output[stp] = lake_state[stp].index.get_level_values('dcm')
assignment_output['mp'] = [x if str(y)=='nan' else y for x, y in \
        zip(assignment_output['step0'], assignment_output['step5'])]
assignment_output['step'] = [0 if str(x) == 'nan' else x for x in \
        assignment_output['step']]

area_sum = tracking.join(dca_info, on='dca')[['to', 'step', 'area_sqmi']]
area_sum.replace(dcm_dict, inplace=True)
area_sum = area_sum[1:]
area_sum = area_sum.groupby(['to', 'step']).sum()

mi = pd.MultiIndex.from_product([set(dcm_dict), range(1, 6)], \
        names=['to', 'step'])
summary = pd.DataFrame(index=mi)
summary = summary.join(area_sum)
summary.fillna(0, inplace=True)
summary = summary.unstack('step')
tot = summary.sum().rename('total')
summary = summary.append(tot)
summary.drop('None', inplace=True)

#summary_melt = pd.melt(assignment_output, id_vars=['area_sqmi'], \
#        value_vars=['step'+str(i) for i in range(0, 6)], \
#        var_name='step', value_name='dcm')
#summary_melt.replace(dcm_dict, inplace=True)
#summary = summary_melt.groupby(['dcm', 'step'])['area_sqmi'].agg('sum').unstack()

# write results into output workbook
ws = wb['MP_new']
# write DCA/DCM assignments 
for i in range(0, len(assignment_output), 1):
    offset = 22
    ws.cell(row=i+offset, column=8).value = assignment_output['mp'][i]
    ws.cell(row=i+offset, column=9).value = assignment_output['step'][i]
# write habitat areas and water use
rw = 3
col_ind = {'bw':2, 'mw':3, 'pl':4, 'ms':5, 'md':6, 'water':7}
for j in ['base', 'step0', 'step1', 'step2', 'step3', 'step4', 'step5']:
    for k in col_ind.keys():
        ws.cell(row=rw, column=col_ind[k]).value = total[j][k]
    rw += 1
# write area summary tables
ws = wb['Area Summary']
for i in range(0, len(summary), 1):
    ws.cell(row=i+5, column=1).value = summary.index.get_level_values('to').tolist()[i]
    for j in range(1, 6):
        ws.cell(row=i+5, column=j+2).value = summary.iloc[i, j-1].round(3)
ws = wb['Change Tracking']
row = 1
for r in dataframe_to_rows(tracking, index=True, header=True):
    for i in range(0, len(r)):
        ws.cell(row=row, column=i+1).value = r[i]
    row += 1
wb.save(output_excel)
book = load_workbook(filename=output_excel)
writer = pd.ExcelWriter(output_excel, engine = 'openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
writer.save()
